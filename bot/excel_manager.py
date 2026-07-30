import json
import os
import random
from datetime import datetime
import asyncio
import openpyxl
from bot.config import (
    EXCEL_PATH,
    DEFAULT_CANH_GIOI,
    DEFAULT_EXP,
    DEFAULT_LINH_THACH,
    DEFAULT_HP,
    DEFAULT_MANA,
    DEFAULT_LINH_CAN,
    LINH_CAN_TYPES,
    REALMS
)
from bot.logger import logger

COLUMNS = [
    "DiscordID",
    "Username",
    "Tên",
    "Cảnh giới",
    "EXP",
    "Linh thạch",
    "Linh căn",
    "HP",
    "Mana",
    "Ngày điểm danh",
    "Túi đồ",
    "Buff đột phá"
]

class ExcelManager:
    def __init__(self, file_path: str = EXCEL_PATH):
        self.file_path = file_path
        self.lock = asyncio.Lock()
        self._ensure_excel_exists_sync()

    def _ensure_excel_exists_sync(self):
        """Synchronously create data directory and Excel file if not exists."""
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        if not os.path.exists(self.file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TuSi"
            ws.append(COLUMNS)
            wb.save(self.file_path)
            wb.close()
            logger.info(f"Excel Saved: Initialized new Excel file at {self.file_path}")

    async def get_all_players(self) -> list[dict]:
        """Reads all player rows from Excel and returns as list of dicts."""
        async with self.lock:
            return self._read_all_sync()

    def _read_all_sync(self) -> list[dict]:
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active
        players = []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            return []

        header = [str(cell) for cell in rows[0]]
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            player = {}
            for col_name, val in zip(header, row):
                player[col_name] = "" if val is None else str(val)
            
            # Cast numeric types
            try:
                player["EXP"] = int(player.get("EXP", 0))
            except ValueError:
                player["EXP"] = 0
                
            try:
                player["Linh thạch"] = int(player.get("Linh thạch", 0))
            except ValueError:
                player["Linh thạch"] = 0

            try:
                player["HP"] = int(player.get("HP", 100))
            except ValueError:
                player["HP"] = 100

            try:
                player["Mana"] = int(player.get("Mana", 100))
            except ValueError:
                player["Mana"] = 100

            players.append(player)
        wb.close()
        return players

    async def get_player(self, discord_id: str) -> dict | None:
        """Get player dict by Discord ID."""
        players = await self.get_all_players()
        for p in players:
            if str(p.get("DiscordID")) == str(discord_id):
                return p
        return None

    async def get_or_create_player(self, discord_id: str, username: str) -> dict:
        """Gets existing player or creates new row in Excel."""
        async with self.lock:
            players = self._read_all_sync()
            discord_id_str = str(discord_id)

            for p in players:
                if str(p.get("DiscordID")) == discord_id_str:
                    return p

            # Randomize Linh Căn for new disciples
            assigned_linh_can = random.choice(LINH_CAN_TYPES)

            # Create new player
            new_player = {
                "DiscordID": discord_id_str,
                "Username": username,
                "Tên": username,
                "Cảnh giới": DEFAULT_CANH_GIOI,
                "EXP": DEFAULT_EXP,
                "Linh thạch": DEFAULT_LINH_THACH,
                "Linh căn": assigned_linh_can,
                "HP": DEFAULT_HP,
                "Mana": DEFAULT_MANA,
                "Ngày điểm danh": "",
                "Túi đồ": "{}",
                "Buff đột phá": 0
            }

            # Save to Excel
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            
            # Ensure headers match
            header = [cell.value for cell in ws[1]]
            for col in COLUMNS:
                if col not in header:
                    ws.cell(row=1, column=len(header)+1, value=col)
                    header.append(col)

            row_data = []
            for col in header:
                row_data.append(new_player.get(col, ""))

            ws.append(row_data)
            wb.save(self.file_path)
            wb.close()

            logger.info(f"Excel Saved: Registered new disciple {username} ({discord_id_str})")
            return new_player

    async def get_inventory(self, discord_id: str) -> dict:
        player = await self.get_player(discord_id)
        if not player:
            return {}
        inv_str = player.get("Túi đồ", "{}")
        try:
            return json.loads(inv_str) if inv_str else {}
        except Exception:
            return {}

    async def add_item(self, discord_id: str, item_name: str, count: int = 1) -> dict:
        inventory = await self.get_inventory(discord_id)
        inventory[item_name] = inventory.get(item_name, 0) + count
        if inventory[item_name] <= 0:
            inventory.pop(item_name, None)
        await self.update_player(discord_id, {"Túi đồ": json.dumps(inventory, ensure_ascii=False)})
        return inventory

    async def remove_item(self, discord_id: str, item_name: str, count: int = 1) -> tuple[bool, dict]:
        inventory = await self.get_inventory(discord_id)
        current = inventory.get(item_name, 0)
        if current < count:
            return False, inventory
        inventory[item_name] = current - count
        if inventory[item_name] <= 0:
            inventory.pop(item_name, None)
        await self.update_player(discord_id, {"Túi đồ": json.dumps(inventory, ensure_ascii=False)})
        return True, inventory

    async def update_player(self, discord_id: str, updates: dict) -> dict | None:
        """Update specific fields for a player and immediately save to Excel."""
        async with self.lock:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=False))
            if not rows:
                wb.close()
                return None

            header = [cell.value for cell in rows[0]]
            discord_id_col = header.index("DiscordID") if "DiscordID" in header else 0

            target_row_idx = None
            for idx, row in enumerate(rows[1:], start=2):
                if row[discord_id_col].value is not None and str(row[discord_id_col].value) == str(discord_id):
                    target_row_idx = idx
                    break

            if target_row_idx is None:
                wb.close()
                return None

            # Apply updates
            for key, val in updates.items():
                if key in header:
                    col_idx = header.index(key) + 1
                    ws.cell(row=target_row_idx, column=col_idx, value=val)

            wb.save(self.file_path)
            wb.close()
            logger.info(f"Excel Saved: Updated stats for DiscordID {discord_id}")

        return await self.get_player(discord_id)

    async def add_exp(self, discord_id: str, exp_amount: int) -> tuple[dict, int]:
        """Add EXP to player."""
        player = await self.get_player(discord_id)
        if not player:
            return None, 0
        current_exp = int(player.get("EXP", 0))
        new_exp = max(0, current_exp + exp_amount)
        updated = await self.update_player(discord_id, {"EXP": new_exp})
        return updated, exp_amount

    async def add_linh_thach(self, discord_id: str, amount: int) -> tuple[dict, int]:
        """Add or deduct Linh thạch."""
        player = await self.get_player(discord_id)
        if not player:
            return None, 0
        current_lt = int(player.get("Linh thạch", 0))
        new_lt = max(0, current_lt + amount)
        updated = await self.update_player(discord_id, {"Linh thạch": new_lt})
        return updated, amount

    async def check_in(self, discord_id: str) -> tuple[bool, dict, str]:
        """
        Daily check-in logic.
        Returns (success: bool, updated_player: dict, message: str).
        """
        player = await self.get_player(discord_id)
        if not player:
            return False, {}, "Chưa tìm thấy hồ sơ đệ tử."

        today_str = datetime.now().strftime("%Y-%m-%d")
        last_checkin = player.get("Ngày điểm danh", "")

        if last_checkin == today_str:
            return False, player, "Hôm nay đệ tử đã điểm danh rồi, ngày mai hãy quay lại!"

        current_lt = int(player.get("Linh thạch", 0))
        reward = 100
        new_lt = current_lt + reward

        updated = await self.update_player(discord_id, {
            "Linh thạch": new_lt,
            "Ngày điểm danh": today_str
        })

        return True, updated, f"Điểm danh thành công! Nhận được +{reward} Linh Thạch. 💎"

    async def get_top_cultivators(self, limit: int = 10) -> list[dict]:
        """Sort cultivators by realm index and total EXP."""
        players = await self.get_all_players()
        realm_names = [r["name"] for r in REALMS]

        def player_rank_key(p):
            realm = p.get("Cảnh giới", "")
            realm_idx = realm_names.index(realm) if realm in realm_names else -1
            exp = int(p.get("EXP", 0))
            lt = int(p.get("Linh thạch", 0))
            return (realm_idx, exp, lt)

        sorted_players = sorted(players, key=player_rank_key, reverse=True)
        return sorted_players[:limit]
