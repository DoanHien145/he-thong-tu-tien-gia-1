import os
import sqlite3
import asyncio
import time
from datetime import datetime
from bot.logger import logger
from bot.config import (
    DEFAULT_CANH_GIOI,
    DEFAULT_EXP,
    DEFAULT_LINH_THACH,
    DEFAULT_HP,
    DEFAULT_MANA,
    DEFAULT_LINH_CAN,
    LINH_CAN_TYPES,
    REALMS
)

class DatabaseManager:
    """
    SQLite Async-wrapped Database Manager for Discord Cultivation Bot.
    Guarantees persistence across bot restarts and deployments.
    """
    def __init__(self, db_path: str = "data/cultivation.db"):
        self.db_path = db_path
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        self._init_db()

    def _get_connection(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self):
        """Initializes tables and indexes if not existing."""
        conn = self._get_connection()
        cursor = conn.cursor()

        # Players table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS players (
                discord_id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                canh_gioi TEXT NOT NULL,
                exp INTEGER NOT NULL DEFAULT 0,
                linh_thach INTEGER NOT NULL DEFAULT 100,
                hp INTEGER NOT NULL DEFAULT 100,
                mana INTEGER NOT NULL DEFAULT 100,
                linh_can TEXT NOT NULL,
                ngay_diem_danh TEXT DEFAULT '',
                buff_dot_pha INTEGER DEFAULT 0,
                cooldown_tu_luyen REAL DEFAULT 0,
                cooldown_sukien REAL DEFAULT 0,
                song_tu_partner TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)

        # Inventory table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS inventory (
                discord_id TEXT NOT NULL,
                item_name TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (discord_id, item_name)
            )
        """)

        # World Boss table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS world_boss (
                date TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                hp INTEGER NOT NULL,
                max_hp INTEGER NOT NULL,
                is_dead INTEGER DEFAULT 0
            )
        """)

        # Boss Damage Leaderboard table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS boss_damage (
                date TEXT NOT NULL,
                discord_id TEXT NOT NULL,
                name TEXT NOT NULL,
                damage INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (date, discord_id)
            )
        """)

        # Quests Claimed Log table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS quests_claimed (
                date TEXT NOT NULL,
                discord_id TEXT NOT NULL,
                quest_id TEXT NOT NULL,
                PRIMARY KEY (date, discord_id, quest_id)
            )
        """)

        conn.commit()

        # Check if database is empty (e.g., after a container redeploy or fresh environment)
        cursor.execute("SELECT COUNT(*) FROM players")
        count = cursor.fetchone()[0]
        conn.close()

        if count == 0:
            logger.info("Cơ sở dữ liệu SQLite trống (có thể do redeploy). Đang tự động khôi phục dữ liệu từ OneDrive/Excel...")
            try:
                from bot.import_onedrive import download_and_import_onedrive
                players_count, items_count = download_and_import_onedrive(db_file=self.db_path)
                logger.info(f"Đã tự động khôi phục thành công {players_count} tu sĩ & {items_count} vật phẩm từ OneDrive!")
            except Exception as e:
                logger.error(f"Không thể tải từ OneDrive, đang thử khôi phục từ file local data.xlsx: {e}")
                try:
                    self._restore_from_local_excel()
                except Exception as ex:
                    logger.error(f"Lỗi khôi phục local excel: {ex}")
        else:
            logger.info(f"SQLite Database đã sẵn sàng ({count} tu sĩ) tại: {self.db_path}")

    def _restore_from_local_excel(self):
        """Fallback to restore SQLite from data/data.xlsx if available."""
        excel_path = "data/data.xlsx"
        if not os.path.exists(excel_path):
            logger.warning("Không tìm thấy file data/data.xlsx để khôi phục.")
            return

        import openpyxl
        wb = openpyxl.load_workbook(excel_path)
        if "Players" not in wb.sheetnames:
            return

        conn = self._get_connection()
        cursor = conn.cursor()
        ws_players = wb["Players"]

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = list(ws_players.iter_rows(values_only=True))
        if len(rows) > 1:
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                discord_id = str(row[0]).strip()
                ten = str(row[1]).strip() if len(row) > 1 and row[1] else "Tu sĩ"
                canh_gioi = str(row[2]).strip() if len(row) > 2 and row[2] else "Luyện Khí tầng 1"
                try: exp = int(row[3])
                except: exp = 0
                try: linh_thach = int(row[4])
                except: linh_thach = 100
                try: hp = int(row[5])
                except: hp = 100
                try: mana = int(row[6])
                except: mana = 100
                linh_can = str(row[7]).strip() if len(row) > 7 and row[7] else "Thiên Linh Căn"
                ngay_diem_danh = str(row[8]).strip() if len(row) > 8 and row[8] else ""

                cursor.execute("""
                    INSERT INTO players (
                        discord_id, name, canh_gioi, exp, linh_thach, hp, mana, linh_can, ngay_diem_danh, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(discord_id) DO UPDATE SET
                        name = excluded.name, canh_gioi = excluded.canh_gioi, exp = excluded.exp,
                        linh_thach = excluded.linh_thach, hp = excluded.hp, mana = excluded.mana,
                        linh_can = excluded.linh_can, ngay_diem_danh = excluded.ngay_diem_danh
                """, (discord_id, ten, canh_gioi, exp, linh_thach, hp, mana, linh_can, ngay_diem_danh, created_at))

        if "Inventory" in wb.sheetnames:
            ws_inv = wb["Inventory"]
            inv_rows = list(ws_inv.iter_rows(values_only=True))
            if len(inv_rows) > 1:
                for row in inv_rows[1:]:
                    if not row or not row[0]:
                        continue
                    discord_id = str(row[0]).strip()
                    item_name = str(row[1]).strip() if len(row) > 1 else ""
                    try: qty = int(row[2])
                    except: qty = 0
                    if item_name and qty > 0:
                        cursor.execute("""
                            INSERT INTO inventory (discord_id, item_name, quantity)
                            VALUES (?, ?, ?)
                            ON CONFLICT(discord_id, item_name) DO UPDATE SET quantity = excluded.quantity
                        """, (discord_id, item_name, qty))

        conn.commit()
        conn.close()
        logger.info("Khôi phục thành công dữ liệu từ local data.xlsx vào SQLite!")

    async def get_or_create_player(self, discord_id: str, name: str) -> dict:
        """Fetch player data or auto-register if new."""
        return await asyncio.to_thread(self._sync_get_or_create_player, str(discord_id), name)

    def _sync_get_or_create_player(self, discord_id: str, name: str) -> dict:
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM players WHERE discord_id = ?", (discord_id,))
        row = cursor.fetchone()

        if row:
            player_dict = dict(row)
            # Standardize key names for cogs compatibility
            conn.close()
            return self._format_player_dict(player_dict)

        # Create new player
        import random
        linh_can = random.choice(LINH_CAN_TYPES)
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cursor.execute("""
            INSERT INTO players (
                discord_id, name, canh_gioi, exp, linh_thach, hp, mana, linh_can, ngay_diem_danh, buff_dot_pha, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            discord_id, name, DEFAULT_CANH_GIOI, DEFAULT_EXP, DEFAULT_LINH_THACH,
            DEFAULT_HP, DEFAULT_MANA, linh_can, "", 0, created_at
        ))

        # Starter items
        cursor.execute("""
            INSERT OR REPLACE INTO inventory (discord_id, item_name, quantity)
            VALUES (?, ?, ?)
        """, (discord_id, "Luyện Khí Đan", 2))

        cursor.execute("""
            INSERT OR REPLACE INTO inventory (discord_id, item_name, quantity)
            VALUES (?, ?, ?)
        """, (discord_id, "Tam Diệp Thảo", 3))

        conn.commit()

        cursor.execute("SELECT * FROM players WHERE discord_id = ?", (discord_id,))
        new_row = cursor.fetchone()
        conn.close()

        logger.info(f"Registered new cultivator: {name} (ID: {discord_id}, Linh Căn: {linh_can})")
        return self._format_player_dict(dict(new_row))

    def _format_player_dict(self, row: dict) -> dict:
        """Converts DB row names into friendly Dict keys used in Cogs."""
        return {
            "ID Discord": str(row["discord_id"]),
            "Tên": row["name"],
            "Cảnh giới": row["canh_gioi"],
            "EXP": int(row["exp"]),
            "Linh thạch": int(row["linh_thach"]),
            "HP": int(row["hp"]),
            "Mana": int(row["mana"]),
            "Linh căn": row["linh_can"],
            "Ngày điểm danh": row["ngay_diem_danh"],
            "Buff đột phá": int(row.get("buff_dot_pha", 0)),
            "Cooldown tu luyện": float(row.get("cooldown_tu_luyen", 0)),
            "Cooldown sự kiện": float(row.get("cooldown_sukien", 0)),
            "Song tu partner": row.get("song_tu_partner", "")
        }

    async def get_player(self, discord_id: str) -> dict | None:
        return await asyncio.to_thread(self._sync_get_player, str(discord_id))

    def _sync_get_player(self, discord_id: str) -> dict | None:
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM players WHERE discord_id = ?", (discord_id,))
        row = cursor.fetchone()
        conn.close()
        return self._format_player_dict(dict(row)) if row else None

    async def get_all_players(self) -> list[dict]:
        return await asyncio.to_thread(self._sync_get_all_players)

    def _sync_get_all_players(self) -> list[dict]:
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM players ORDER BY exp DESC")
        rows = cursor.fetchall()
        conn.close()
        return [self._format_player_dict(dict(r)) for r in rows]

    async def update_player(self, discord_id: str, updates: dict) -> dict:
        return await asyncio.to_thread(self._sync_update_player, str(discord_id), updates)

    def _sync_update_player(self, discord_id: str, updates: dict) -> dict:
        mapping = {
            "Tên": "name",
            "Cảnh giới": "canh_gioi",
            "EXP": "exp",
            "Linh thạch": "linh_thach",
            "HP": "hp",
            "Mana": "mana",
            "Linh căn": "linh_can",
            "Ngày điểm danh": "ngay_diem_danh",
            "Buff đột phá": "buff_dot_pha",
            "Cooldown tu luyện": "cooldown_tu_luyen",
            "Cooldown sự kiện": "cooldown_sukien",
            "Song tu partner": "song_tu_partner"
        }

        sql_parts = []
        values = []

        for key, val in updates.items():
            db_col = mapping.get(key, key)
            sql_parts.append(f"{db_col} = ?")
            values.append(val)

        values.append(discord_id)
        sql = f"UPDATE players SET {', '.join(sql_parts)} WHERE discord_id = ?"

        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute(sql, values)
        conn.commit()

        cursor.execute("SELECT * FROM players WHERE discord_id = ?", (discord_id,))
        row = cursor.fetchone()
        conn.close()

        return self._format_player_dict(dict(row))

    async def add_exp(self, discord_id: str, amount: int) -> tuple[dict, bool]:
        """Add EXP to player and check for automatic realm status."""
        return await asyncio.to_thread(self._sync_add_exp, str(discord_id), amount)

    def _sync_add_exp(self, discord_id: str, amount: int) -> tuple[dict, bool]:
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT exp FROM players WHERE discord_id = ?", (discord_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            raise ValueError(f"Player {discord_id} not found")

        new_exp = max(0, int(row["exp"]) + amount)
        cursor.execute("UPDATE players SET exp = ? WHERE discord_id = ?", (new_exp, discord_id))
        conn.commit()

        cursor.execute("SELECT * FROM players WHERE discord_id = ?", (discord_id,))
        updated_row = cursor.fetchone()
        conn.close()

        return self._format_player_dict(dict(updated_row)), False

    async def add_linh_thach(self, discord_id: str, amount: int) -> dict:
        return await asyncio.to_thread(self._sync_add_linh_thach, str(discord_id), amount)

    def _sync_add_linh_thach(self, discord_id: str, amount: int) -> dict:
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT linh_thach FROM players WHERE discord_id = ?", (discord_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            raise ValueError(f"Player {discord_id} not found")

        new_lt = max(0, int(row["linh_thach"]) + amount)
        cursor.execute("UPDATE players SET linh_thach = ? WHERE discord_id = ?", (new_lt, discord_id))
        conn.commit()

        cursor.execute("SELECT * FROM players WHERE discord_id = ?", (discord_id,))
        updated_row = cursor.fetchone()
        conn.close()
        return self._format_player_dict(dict(updated_row))

    async def check_in(self, discord_id: str) -> tuple[bool, dict, str]:
        """Daily check-in reward."""
        return await asyncio.to_thread(self._sync_check_in, str(discord_id))

    def _sync_check_in(self, discord_id: str) -> tuple[bool, dict, str]:
        today = datetime.now().strftime("%Y-%m-%d")
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM players WHERE discord_id = ?", (discord_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return False, {}, "Người chơi chưa được khởi tạo!"

        last_checkin = row["ngay_diem_danh"]
        if last_checkin == today:
            conn.close()
            return False, self._format_player_dict(dict(row)), "Hôm nay bạn đã điểm danh rồi! Hãy quay lại vào ngày mai."

        new_lt = int(row["linh_thach"]) + 150
        new_exp = int(row["exp"]) + 100
        cursor.execute("""
            UPDATE players SET ngay_diem_danh = ?, linh_thach = ?, exp = ? WHERE discord_id = ?
        """, (today, new_lt, new_exp, discord_id))

        # Add bonus daily herb
        cursor.execute("""
            INSERT INTO inventory (discord_id, item_name, quantity)
            VALUES (?, 'Tam Diệp Thảo', 1)
            ON CONFLICT(discord_id, item_name) DO UPDATE SET quantity = quantity + 1
        """, (discord_id,))

        conn.commit()

        cursor.execute("SELECT * FROM players WHERE discord_id = ?", (discord_id,))
        updated_row = cursor.fetchone()
        conn.close()

        msg = f"🎉 Nhận được **+150 Linh Thạch**, **+100 EXP** và **1x Tam Diệp Thảo** quà báo danh!"
        return True, self._format_player_dict(dict(updated_row)), msg

    async def get_inventory(self, discord_id: str) -> dict[str, int]:
        return await asyncio.to_thread(self._sync_get_inventory, str(discord_id))

    def _sync_get_inventory(self, discord_id: str) -> dict[str, int]:
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT item_name, quantity FROM inventory WHERE discord_id = ? AND quantity > 0", (discord_id,))
        rows = cursor.fetchall()
        conn.close()
        return {r["item_name"]: int(r["quantity"]) for r in rows}

    async def add_item(self, discord_id: str, item_name: str, quantity: int) -> dict[str, int]:
        return await asyncio.to_thread(self._sync_add_item, str(discord_id), item_name, quantity)

    def _sync_add_item(self, discord_id: str, item_name: str, quantity: int) -> dict[str, int]:
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO inventory (discord_id, item_name, quantity)
            VALUES (?, ?, ?)
            ON CONFLICT(discord_id, item_name) DO UPDATE SET quantity = max(0, quantity + ?)
        """, (discord_id, item_name, quantity, quantity))
        conn.commit()
        conn.close()
        return self._sync_get_inventory(discord_id)

    async def use_item(self, discord_id: str, item_name: str, quantity: int = 1) -> tuple[bool, str]:
        return await asyncio.to_thread(self._sync_use_item, str(discord_id), item_name, quantity)

    def _sync_use_item(self, discord_id: str, item_name: str, quantity: int = 1) -> tuple[bool, str]:
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT quantity FROM inventory WHERE discord_id = ? AND item_name = ?", (discord_id, item_name))
        row = cursor.fetchone()

        if not row or int(row["quantity"]) < quantity:
            conn.close()
            return False, f"Bạn không có đủ **{item_name}** trong túi đồ!"

        new_qty = int(row["quantity"]) - quantity
        if new_qty > 0:
            cursor.execute("UPDATE inventory SET quantity = ? WHERE discord_id = ? AND item_name = ?", (new_qty, discord_id, item_name))
        else:
            cursor.execute("DELETE FROM inventory WHERE discord_id = ? AND item_name = ?", (discord_id, item_name))

        conn.commit()
        conn.close()
        return True, f"Đã sử dụng thành công **{quantity}x {item_name}**!"

    async def get_top_players(self, limit: int = 10) -> list[dict]:
        return await asyncio.to_thread(self._sync_get_top_players, limit)

    async def get_top_cultivators(self, limit: int = 10) -> list[dict]:
        return await self.get_top_players(limit)

    def _sync_get_top_players(self, limit: int = 10) -> list[dict]:
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM players ORDER BY exp DESC LIMIT ?", (limit,))
        rows = cursor.fetchall()
        conn.close()
        return [self._format_player_dict(dict(r)) for r in rows]

    # --- World Boss Methods ---
    async def get_daily_boss(self) -> dict:
        return await asyncio.to_thread(self._sync_get_daily_boss)

    def _sync_get_daily_boss(self) -> dict:
        today = datetime.now().strftime("%Y-%m-%d")
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM world_boss WHERE date = ?", (today,))
        row = cursor.fetchone()

        if not row:
            import random
            boss_names = [
                "👹 Thượng Cổ Thiên Ma (Viễn Cổ)",
                "🐉 Hắc Ngục Cửu Đầu Long",
                "🦁 Vạn Niên Huyết Kỳ Lân",
                "💀 Vong Hồn U Minh Ma Đế",
                "🔥 Vực Thẫm Viêm Ma Vương"
            ]
            name = random.choice(boss_names)
            max_hp = 366769
            cursor.execute("""
                INSERT INTO world_boss (date, name, hp, max_hp, is_dead)
                VALUES (?, ?, ?, ?, 0)
            """, (today, name, max_hp, max_hp))
            conn.commit()
            cursor.execute("SELECT * FROM world_boss WHERE date = ?", (today,))
            row = cursor.fetchone()

        conn.close()
        return dict(row)

    async def attack_daily_boss(self, discord_id: str, name: str, damage: int) -> tuple[dict, bool]:
        return await asyncio.to_thread(self._sync_attack_daily_boss, str(discord_id), name, damage)

    def _sync_attack_daily_boss(self, discord_id: str, name: str, damage: int) -> tuple[dict, bool]:
        today = datetime.now().strftime("%Y-%m-%d")
        boss = self._sync_get_daily_boss()

        conn = self._get_connection()
        cursor = conn.cursor()

        new_hp = boss["max_hp"]
        is_dead = 0

        cursor.execute("UPDATE world_boss SET hp = ?, is_dead = ? WHERE date = ?", (new_hp, is_dead, today))

        # Update leaderboard
        cursor.execute("""
            INSERT INTO boss_damage (date, discord_id, name, damage)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(date, discord_id) DO UPDATE SET damage = damage + ?, name = ?
        """, (today, discord_id, name, damage, damage, name))

        conn.commit()

        cursor.execute("SELECT * FROM world_boss WHERE date = ?", (today,))
        updated_boss = dict(cursor.fetchone())
        conn.close()

        return updated_boss, (is_dead == 1 and boss["is_dead"] == 0)

    async def get_boss_leaderboard(self, limit: int = 10) -> list[dict]:
        return await asyncio.to_thread(self._sync_get_boss_leaderboard, limit)

    def _sync_get_boss_leaderboard(self, limit: int = 10) -> list[dict]:
        today = datetime.now().strftime("%Y-%m-%d")
        conn = self._get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT name, damage, discord_id FROM boss_damage
            WHERE date = ? ORDER BY damage DESC LIMIT ?
        """, (today, limit))

        rows = cursor.fetchall()
        conn.close()
        return [dict(r) for r in rows]

    async def export_to_excel(self, excel_path: str = "data/data.xlsx"):
        """Generates dynamic Excel dump from SQLite for download endpoints."""
        await asyncio.to_thread(self._sync_export_to_excel, excel_path)

    def _sync_export_to_excel(self, excel_path: str = "data/data.xlsx"):
        import openpyxl
        wb = openpyxl.Workbook()

        # Sheet 1: Players
        ws_players = wb.active
        ws_players.title = "Players"
        ws_players.append([
            "ID Discord", "Tên", "Cảnh giới", "EXP", "Linh thạch", "HP", "Mana", "Linh căn", "Ngày điểm danh"
        ])

        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM players")
        for r in cursor.fetchall():
            ws_players.append([
                r["discord_id"], r["name"], r["canh_gioi"], r["exp"], r["linh_thach"],
                r["hp"], r["mana"], r["linh_can"], r["ngay_diem_danh"]
            ])

        # Sheet 2: Inventory
        ws_inv = wb.create_sheet(title="Inventory")
        ws_inv.append(["ID Discord", "Vật phẩm", "Số lượng"])

        cursor.execute("SELECT * FROM inventory WHERE quantity > 0")
        for r in cursor.fetchall():
            ws_inv.append([r["discord_id"], r["item_name"], r["quantity"]])

        conn.close()
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb.save(excel_path)

    async def save(self):
        """Export current SQLite state to data/data.xlsx."""
        await self.export_to_excel()
